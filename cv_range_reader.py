#!/usr/bin/env python3
# coding: utf-8
"""
excelファイルを読み取ってCVの測定電位を抽出する。
"""

import os
import sys
from datetime import date
from pprint import pprint
# from subprocess import Popen

import openpyxl


def read_xlsx(filepath):
    """
    xlsxファイルから電位情報を読み取る
    filepath: xlsxファイルのパス
    """
    # ファイルを開く
    wb = openpyxl.load_workbook(filepath)

    # ブック名
    bookname = os.path.splitext(os.path.basename(filepath))[0]
    # シート一覧
    sheetnames = wb.sheetnames

    print('bookname  :', bookname)
    print('sheetnames:', sheetnames)

    # 電位リスト [[Emin, Emax], ...]
    l = []

    for sheet in wb:
        d = {'bookname': bookname, 'sheetname': sheet.title, 'Emin': None, 'Emax': None}
        # 電位の列であることを確認
        if sheet['A1'].value == 'E /V':
            # 列の値を取得
            column = [cell.value for cell in sheet['A']]
            # 最小値と最大値を取得
            d['Emin'], d['Emax'] = min(column[1:]), max(column[1:])

        l.append(d)

    return l


def make_xlsx(l, savepath):
    """
    受け取った「辞書のリスト」をxlsxファイルに書き出す
    l       : 辞書のリスト
    savepath: ファイルを保存するパス
    """
    try:
        # 既存ブック
        wb = openpyxl.load_workbook(savepath)
        # 新規シート
        sheet = wb.create_sheet()
    except Exception as e:
        # 新規ブック
        print(e)
        wb = openpyxl.Workbook()
        sheet = wb.active
    sheet['A1'].value = 'bookname'
    sheet['B1'].value = 'sheetname'
    sheet['C1'].value = 'Emin'
    sheet['D1'].value = 'Emax'
    sheet['E1'].value = 'update'

    for i, d in enumerate(l):
        # print('i={}: {}'.format(i, d))
        sheet.cell(row=i + 2, column=1).value = d['bookname']
        sheet.cell(row=i + 2, column=2).value = d['sheetname']
        sheet.cell(row=i + 2, column=3).value = d['Emin']
        sheet.cell(row=i + 2, column=4).value = d['Emax']
        sheet.cell(row=i + 2, column=5).value = str(date.today())  # 実行日

    # ファイル出力
    wb.save(savepath)


def main():
    """
    全体の処理
    """
    # 読み取るファイルを指定
    print('ファイルを指定してください。')
    filepath = input('> ').strip('"')

    if not filepath.endswith(('.xlsx', '.xlsm')):
        filepath += '.xlsx'

    try:
        l = read_xlsx(filepath)
    except FileNotFoundError as e:
        print(e)
        print('指定したファイルがありません。')
        sys.exit('終了します。')

    print('\n取得結果----------------')
    pprint(l)
    print('------------------------\n')

    savepath = 'result(CVRangeReader).xlsx'
    make_xlsx(l, savepath)

    print('できた')

    # つくったファイルを開く
    # Popen(['start', savepath], shell=True)

if __name__ == '__main__':
    while True:
        main()
